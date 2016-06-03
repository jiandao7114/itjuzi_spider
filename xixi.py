#-*- coding: UTF-8 -*-

import sys
import time
import urllib2
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook

reload(sys)
sys.setdefaultencoding('utf8')


#Some User Agents
hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
{'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


def http_connet(url,try_times=0):
    try:
        req = urllib2.Request(url, headers=hds[np.random.randint(0,len(hds))])
        source_code = urllib2.urlopen(req,timeout=10).read()
        plain_text=str(source_code)
    except Exception, e:
        # print e
        if try_times<3:
            try_times+=1
            print '网络异常，重试第 '+str(try_times)+' 次'
            return http_connet(url,try_times)
    soup = BeautifulSoup(plain_text,"html.parser")
    return soup


def itjuzi_spider():
    global PAGE_NUM
    PAGE_NUM=1
    try_times=0
    while(PAGE_NUM < 1001):

        project_list=[]
        
        time.sleep(3) # 过于频繁，可能导致IP地址被禁

        print '开始爬取第 %d 页数据' % PAGE_NUM
        url='https://www.itjuzi.com/investevents?page='+ str(PAGE_NUM)
        soup = http_connet(url)
        list_soup = soup.findAll('ul', {'class': 'list-main-eventset'})
        
        try_times+=1;
        if list_soup==None and try_times<10:
            continue
        elif list_soup==None or len(list_soup)<=1:
            break # 10次无数据后，停止查询

        for project_info in list_soup[1].findAll('li'):
            try:
                content = project_info.find('i', {'class':'cell pic'}).a['href']
                content=get_content(content)
            except:
                content =''

            try:
                times = project_info.find('i', {'class':'cell round'}).span.string.strip()
            except:
                times = '时间:保密'

            try:
                title = project_info.find('p',{'class':'title'}).a.span.string.strip()
            except:
                title = '标题:缺失'

            try:
                ptype = project_info.find('span',{'class':'tags t-small c-gray-aset'}).a.string.strip()
            except:
                ptype = '类别:缺失'

            try:
                city = project_info.find('span',{'class':'loca c-gray-aset t-small'}).a.string.strip()
            except:
                city = '城市:缺失'

            try:
                lunci = project_info.find('span', {'class':'tag gray'}).string.strip()
            except:
                lunci = '轮次:缺失'

            try:
                money = project_info.find('i', {'class':'cell fina'}).string.strip()
            except:
                money = '融资额:缺失'

            try:
                touzis=''
                investorset = project_info.find('span',{'class':'investorset'})
                if investorset.a!=None:
                    touzis+=investorset.a.string.strip()+' '
                touzia = investorset.findAll('span', {'class':'c-gray'})
                if touzia != None:
                    for touzi in touzia:
                        touzis+=touzi.string.strip()+' '
            except:
                touzis = '领头方:缺失'

            project_list.append([times,title,ptype,city,lunci,money,touzis,content])
            try_times=0 #set 0 when got valid information
        print '第 %d 页数据下载完毕' % PAGE_NUM
        write_excel(project_list)
        PAGE_NUM+=1

def get_content(url):
    soup = http_connet(url)
    content=soup.find('div',{'class':'block'}).hr.find_next('p').string.strip()
    return content

def write_excel(lists):
    wb = load_workbook(filename=r'IT橘子项目融资爬取.xlsx')
    ws = wb['项目融资数据']
    count=1
    for info in lists:
        ws.append([(PAGE_NUM-1)*10+count,info[1],info[7],info[2],info[3],info[4],info[5],info[0],info[6]])
        count+=1
    print '插入'+str(count-1)+'行数据'
    wb.save('IT橘子项目融资爬取.xlsx')

def create_excel():
    wb=Workbook()
    ws=wb.active
    ws.title='项目融资数据'.decode()
    ws.append(['序号','项目名称','项目简介','行业','城市','轮次','融资金额','投资时间','领投方'])
    wb.save('IT橘子项目融资爬取.xlsx')

if __name__=='__main__':
    # create_excel()
    itjuzi_spider()


